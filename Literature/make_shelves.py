from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SHELVES = {
    1: "RURO / Latent Jobs / Constrained Labour Supply",
    2: "Welfare Measurement for Discrete Choice",
    3: "Equivalent Income and Beyond-GDP Welfare",
    4: "Responsibility, Compensation and Fairness",
    5: "Inequality of Opportunity and Decomposition",
    6: "Microsimulation and Tax-Benefit Policy Evaluation",
}

SHELF_COLORS = {
    1: "D9EAF7",
    2: "D9F0E8",
    3: "FFF2CC",
    4: "FCE4D6",
    5: "EAD1DC",
    6: "E2EFDA",
}

PAPERS = [
    # Shelf 1
    ("Aaberge Colombino Strom 1999",  "Aaberge, Colombino & Strom (1999) - Labour Supply in Italy", 1),
    ("Aaberge Colombino Strom 2000",  "Aaberge, Colombino & Strom (2000) - Flat Tax: Italy, Norway, Sweden", 1),
    ("Aaberge Colombino Strom 2004",  "Aaberge, Colombino & Strom (2004) - Do More Equal Slices Shrink the Cake?", 1),
    ("Aaberge Colombino 2006",        "Aaberge & Colombino (2006) - Designing Optimal Taxes", 1),
    ("Aaberge Colombino 2011",        "Aaberge & Colombino (2011) - Empirical Optimal Income Taxation (Norway)", 1),
    ("Aaberge Colombino 2012",        "Aaberge & Colombino (2012) - Family Background and Optimal Taxes", 1),
    ("Aaberge Colombino 2013",        "Aaberge & Colombino (2013) - Optimal Income Taxes via Labour Supply Model", 1),
    ("Aaberge Colombino 2018",        "Aaberge & Colombino (2018) - Structural Labour Supply and Microsimulation", 1),
    ("Aaberge Colombino Wennemo 2009","Aaberge, Colombino & Wennemo (2009) - Evaluating Choice Sets", 1),
    ("Aaberge Dagsvik Strom 1995",    "Aaberge, Dagsvik & Strom (1995) - Labor Supply Responses and Welfare Effects of Tax Reforms", 1),
    ("Bargain Doorley 2013",          "Bargain & Doorley (2013) - RD Design and Youth Inactivity", 1),
    ("Beffy et al 2019",              "Beffy et al. (2019) - Labour Supply with Restricted Choices", 1),
    ("Bloemen 2000",                  "Bloemen (2000) - Labour Supply with Job Offer Restrictions", 1),
    ("Bloemen 2008",                  "Bloemen (2008) - Job Search and Hours Restrictions", 1),
    ("Capeau Decoster 2016",          "Capeau & Decoster (2016) - Re-tiring and Job Opportunities", 1),
    ("Capeau Decoster Dekkers 2016",  "Capeau, Decoster & Dekkers (2016) - RURO Model (Belgium)", 1),
    ("Dagsvik Jia 2016",              "Dagsvik & Jia (2016) - Latent Jobs: Identification", 1),
    ("Dagsvik et al 2014",            "Dagsvik et al. (2014) - Modeling Labor Supply as Latent Jobs", 1),
    ("Dagsvik Strom 2006",            "Dagsvik & Strom (2006) - Sectoral Labour Supply", 1),
    ("Loffler Peichl Siegloch 2014",  "Loffler, Peichl & Siegloch (2014) - Wage Exogeneity", 1),
    ("Loffler Peichl Siegloch 2018",  "Loffler, Peichl & Siegloch (2018) - Sensitivity of Labour Supply Estimates", 1),
    ("Peichl Siegloch 2012",          "Peichl & Siegloch (2012) - Labor Demand Effects", 1),
    ("Van Soest 1995",                "Van Soest (1995) - Structural Models of Family Labor Supply", 1),
    # Shelf 2
    ("Bargain et al 2013 welfare",    "Bargain et al. (2013) - Welfare, Labor Supply and Heterogeneous Preferences", 2),
    ("Bhattacharya 2015",             "Bhattacharya (2015) - Nonparametric Welfare Analysis for Discrete Choice", 2),
    ("Bhattacharya 2018",             "Bhattacharya (2018) - Empirical Welfare Analysis for Discrete Choice", 2),
    ("Blomquist Newey 2002",          "Blomquist & Newey (2002) - Nonparametric Estimation with Nonlinear Budget Sets", 2),
    ("Capeau De Sadeleer Maes 2021",  "Capeau et al. (2021) - Nonparametric Welfare (Levels and Differences)", 2),
    ("Carpantier Sapata 2016",        "Carpantier & Sapata (2016) - Empirical Welfare When Preferences Matter", 2),
    ("Chetty 2009",                   "Chetty (2009) - Sufficient Statistics for Welfare Analysis", 2),
    ("Dagsvik Karlstrom 2005",        "Dagsvik & Karlstrom (2005) - Compensating Variation in Random Utility", 2),
    ("Mas Pallais 2017",              "Mas & Pallais (2017) - Valuing Alternative Work Arrangements", 2),
    # Shelf 3
    ("Bargain et al 2014 inequality", "Bargain et al. (2014) - Inequality Aversion across Countries", 3),
    ("Bargain Peichl 2016",           "Bargain & Peichl (2016) - Own-wage Labor Supply Elasticities", 3),
    ("Becker Philipson Soares 2005",  "Becker, Philipson & Soares (2005) - Quantity and Quality of Life", 3),
    ("Decancq Fleurbaey Schokkaert 2015", "Decancq, Fleurbaey & Schokkaert (2015) - Happiness and Equivalent Incomes", 3),
    ("Decancq Schokkaert 2016",       "Decancq & Schokkaert (2016) - Beyond GDP (Equivalent Incomes)", 3),
    ("Durand 2015",                   "Durand (2015) - OECD Better Life Initiative", 3),
    ("Fleurbaey 2009",                "Fleurbaey (2009) - Beyond GDP", 3),
    ("Fleurbaey Gaulier 2009",        "Fleurbaey & Gaulier (2009) - International Comparisons by Equivalent Incomes", 3),
    ("Fleurbaey Maniquet 2017",       "Fleurbaey & Maniquet (2017) - Fairness and Well-being Measurement", 3),
    ("Fleurbaey Maniquet 2018a",      "Fleurbaey & Maniquet (2018a) - Inequality-averse Well-being", 3),
    ("Fleurbaey Maniquet 2019",       "Fleurbaey & Maniquet (2019) - Non-classical Goods", 3),
    ("Jones Klenow 2016",             "Jones & Klenow (2016) - Beyond GDP? Welfare across Countries", 3),
    # Shelf 4
    ("Chone Laroque 2010",            "Chone & Laroque (2010) - Negative Marginal Tax Rates and Heterogeneity", 4),
    ("Fleurbaey 1995a",               "Fleurbaey (1995a) - Equal Opportunity or Equal Social Outcome?", 4),
    ("Fleurbaey 1995b",               "Fleurbaey (1995b) - Three Solutions for the Compensation Problem", 4),
    ("Fleurbaey Maniquet 2006",       "Fleurbaey & Maniquet (2006) - Fair Income Tax", 4),
    ("Fleurbaey Maniquet 2007",       "Fleurbaey & Maniquet (2007) - Help the Low Skilled or the Hardworking", 4),
    ("Fleurbaey Maniquet 2011 ch10",  "Fleurbaey & Maniquet (2011) - Ch. 10: Unequal Skills", 4),
    ("Fleurbaey Maniquet 2011 ch11",  "Fleurbaey & Maniquet (2011) - Ch. 11: Income Taxation", 4),
    ("Fleurbaey Maniquet 2018b",      "Fleurbaey & Maniquet (2018b) - Optimal Taxation and Fairness", 4),
    ("Jacquet Jia Thoresen 2026",     "Jacquet, Jia & Thoresen (2026) - Responsibility in Fairness Measurement", 4),
    ("Maniquet 2008",                 "Maniquet (2008) - Social Orderings for Indivisible Objects", 4),
    ("Roemer 2002",                   "Roemer (2002) - Equality of Opportunity: A Progress Report", 4),
    ("Roemer Trannoy 2016",           "Roemer & Trannoy (2016) - Equality of Opportunity: Theory and Measurement", 4),
    ("Saez 2002",                     "Saez (2002) - Optimal Income Transfer Programs", 4),
    ("Saez Stantcheva 2016",          "Saez & Stantcheva (2016) - Generalized Social Marginal Welfare Weights", 4),
    ("Sher 2024",                     "Sher (2024) - Inconsistency of GSMWW", 4),
    ("Valletta 2009",                 "Valletta (2009) - Fair Solution to the Compensation Problem", 4),
    # Shelf 5
    ("Audoly et al 2025",             "Audoly et al. (2025) - Shapley-Owen-Shorrocks Decomposition", 5),
    ("Bourguignon Ferreira Menendez 2007", "Bourguignon, Ferreira & Menendez (2007) - Inequality of Opportunity in Brazil", 5),
    ("Brunori Ferreira Peragine 2013","Brunori, Ferreira & Peragine (2013) - IOp across Countries", 5),
    ("Ferreira Gignoux 2011",         "Ferreira & Gignoux (2011) - Measurement of IOp (Latin America)", 5),
    ("Shorrocks 2013",                "Shorrocks (2013) - Shapley Value Decomposition", 5),
    # Shelf 6
    ("Bargain Orsini 2006",           "Bargain & Orsini (2006) - In-work Policies in Europe", 6),
    ("Blundell 2006",                 "Blundell (2006) - Earned Income Tax Credit Policies", 6),
    ("Blundell et al 2000",           "Blundell et al. (2000) - Working Families' Tax Credit", 6),
    ("Bourguignon Spadaro 2006",      "Bourguignon & Spadaro (2006) - Microsimulation and Redistribution", 6),
    ("Immervoll et al 2007",          "Immervoll et al. (2007) - Welfare Reform in Europe", 6),
    ("Sutherland Figari 2013",        "Sutherland & Figari (2013) - EUROMOD", 6),
]

PAPERS_SORTED = sorted(PAPERS, key=lambda x: x[1].lower())

wb = Workbook()
ws = wb.active
ws.title = "Library Shelves"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.column_dimensions["A"].width = 72
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 54
ws.row_dimensions[1].height = 30

headers = ["Paper", "Shelf", "Shelf Name"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

for row_idx, (_, citation, shelf_num) in enumerate(PAPERS_SORTED, start=2):
    fill = PatternFill("solid", fgColor=SHELF_COLORS[shelf_num])
    c1 = ws.cell(row=row_idx, column=1, value=citation)
    c2 = ws.cell(row=row_idx, column=2, value=shelf_num)
    c3 = ws.cell(row=row_idx, column=3, value=SHELVES[shelf_num])
    for c in (c1, c2, c3):
        c.fill = fill
        c.border = border
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(vertical="center")
    c2.alignment = Alignment(horizontal="center", vertical="center")

ws.freeze_panes = "A2"

# Legend sheet
ws2 = wb.create_sheet("Legend")
ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 55
ws2.cell(row=1, column=1, value="Shelf").font = Font(bold=True, size=11)
ws2.cell(row=1, column=2, value="Shelf Name").font = Font(bold=True, size=11)
for i, (num, name) in enumerate(SHELVES.items(), start=2):
    fill = PatternFill("solid", fgColor=SHELF_COLORS[num])
    c1 = ws2.cell(row=i, column=1, value=num)
    c2 = ws2.cell(row=i, column=2, value=name)
    for c in (c1, c2):
        c.fill = fill
        c.border = border
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(vertical="center")
    c1.alignment = Alignment(horizontal="center", vertical="center")

OUTPUT = r"C:\Users\hisham\Desktop\Job_Market_paper\Literature\JMP_Library_Shelves.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}  ({len(PAPERS_SORTED)} papers)")
